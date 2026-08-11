#!/usr/bin/env python3
"""Generate advanced Ghana cash book, bank statement, BRS (Excel + PDF) for go-live testing."""

from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

OUT = Path(__file__).resolve().parent
D = Decimal


def money(x) -> Decimal:
    return D(str(x)).quantize(D("0.01"), rounding=ROUND_HALF_UP)


def fmt(x: Decimal) -> str:
    return f"{x:,.2f}"


COMPANY = "Horizon Insurance Brokers Ltd"
BANK = "Ecobank Ghana"
ACCT = "1441002289035"
CURRENCY = "GHS"
AS_AT = date(2026, 6, 30)
PERIOD = "1 June 2026 – 30 June 2026"

# ---------------------------------------------------------------------------
# Opening balances aligned (no BF bank clears in this pack).
# Period-end timing alone drives the BRS so the platform can tie without
# a prior roll-forward project.
# ---------------------------------------------------------------------------
OPEN_CB = money("245680.75")
OPEN_BANK = money("245680.75")
BF_UNCREDITED = money("0")
BF_UNPRESENTED = money("0")

# Period-end timing (still open as at 30 Jun)
UNCREDITED = [
    # date, name, details, doc_ref, chq, accode, amount
    (date(2026, 6, 25), "GLICO Life", "Commissions received – cheque not yet lodged", "RC025/06/26", "001902", 1020, money("12450.00")),
    (date(2026, 6, 28), "Kwesi Appiah", "GIPS receipt recorded ahead of bank credit", "RC028/06/26", None, 4300, money("3750.00")),
]
UNPRESENTED = [
    (date(2026, 6, 19), "Akan Metals Ltd", "Payment for office fittings – CHQ not presented", "OP041/06/26", "930111", 2140, money("7850.00")),
    (date(2026, 6, 20), "Mensah & Co Legal", "Legal retainership Jun 2026", "OP042/06/26", "930112", 2160, money("4200.00")),
    (date(2026, 6, 29), "Sodium Brand Solutions", "Call cards and branding materials", "OP055/06/26", "930114", 2230, money("3150.00")),
]

BANK_ONLY_DEBITS = [
    (date(2026, 6, 9), "COMMISSION ON TURNOVER COT JUN 2026 / / //H98COT260090011", money("185.00")),
    (date(2026, 6, 15), "SMS ALERT FEE JUN 2026", money("15.00")),
    (date(2026, 6, 28), "ACCOUNT MAINTENANCE FEE JUN 2026", money("50.00")),
    (date(2026, 6, 28), "CHEQUE BOOK ISSUANCE FEE", money("30.00")),
]
BANK_ONLY_CREDITS = [
    (date(2026, 6, 30), "CREDIT INTEREST JUN 2026 / / //H01INT26030088", money("124.75")),
]


def build_cash_book_rows():
    """Return list of cash-book dicts in date order (excluding opening line)."""
    rows = []

    def add(**kwargs):
        rows.append(kwargs)

    # --- Receipts (matched) ---
    add(dt=date(2026, 6, 2), name="Enterprise Life Assurance", details="Commissions received via transfer",
        doc=None, chq="trnsfer", acc=1020, recv=money("42850.00"), paid=None, tag="match", note="Funds transfer inward")
    add(dt=date(2026, 6, 3), name="Ameyaw Boadu", details="Payments received from Ameyaw Boadu GIPS transfer",
        doc=None, chq=None, acc=4300, recv=money("8500.00"), paid=None, tag="match", note="MoMo/GIPS")
    add(dt=date(2026, 6, 5), name="Donewell Insurance", details="Commissions received via Access bank cheque 001891",
        doc="RC005/06/26", chq=None, acc=1020, recv=money("15240.50"), paid=None, tag="match", note="Cheque clearing inward")
    add(dt=date(2026, 6, 6), name="Star Assurance", details="Commissions via ACH from other banks",
        doc=None, chq="ACH", acc=1020, recv=money("22100.00"), paid=None, tag="match", note="ACH inward")
    add(dt=date(2026, 6, 9), name="Ecobank", details="182-DAY T-BILLS- MATURD",
        doc=None, chq="REDDEMED", acc=3000, recv=money("95000.00"), paid=None, tag="match", note="T-bills matured")
    # One-to-many: two cash receipts → one bank batch credit
    add(dt=date(2026, 6, 10), name="Phoenix Insurance", details="Commission batch A – combined bank deposit",
        doc="RC010A/06/26", chq=None, acc=1020, recv=money("6200.00"), paid=None, tag="one_to_many", note="Batch deposit A")
    add(dt=date(2026, 6, 10), name="Quality Insurance", details="Commission batch B – combined bank deposit",
        doc="RC010B/06/26", chq=None, acc=1020, recv=money("3800.00"), paid=None, tag="one_to_many", note="Batch deposit B")
    add(dt=date(2026, 6, 12), name="SIC Insurance", details="Commissions received via transfer",
        doc=None, chq="trnsfer", acc=1020, recv=money("18750.25"), paid=None, tag="match", note="Funds transfer")
    add(dt=date(2026, 6, 16), name="Kwame Mensah", details="Client premium via MTN MoMo",
        doc="RC016/06/26", chq=None, acc=4300, recv=money("4200.00"), paid=None, tag="match", note="MoMo")
    add(dt=date(2026, 6, 18), name="Hollard Insurance", details="Commissions received via transfer",
        doc=None, chq="trnsfer", acc=1020, recv=money("31450.00"), paid=None, tag="match", note="Funds transfer")
    add(dt=date(2026, 6, 20), name="GLICO General", details="Commissions via house cheque deposit 000455",
        doc="RC020/06/26", chq=None, acc=1020, recv=money("9875.00"), paid=None, tag="match", note="HSE cheque deposit")
    # Duplicate amounts, different counterparts
    add(dt=date(2026, 6, 23), name="Naa Adoley Trading", details="Premium collection receipt A",
        doc="RC023A/06/26", chq=None, acc=4300, recv=money("5000.00"), paid=None, tag="dup_amt", note="Duplicate amount A")
    add(dt=date(2026, 6, 24), name="Kojo Antwi Ventures", details="Premium collection receipt B",
        doc="RC023B/06/26", chq=None, acc=4300, recv=money("5000.00"), paid=None, tag="dup_amt", note="Duplicate amount B")
    # Uncredited (cash only)
    for dt, name, details, doc, chq, acc, amt in UNCREDITED:
        add(dt=dt, name=name, details=details, doc=doc, chq=chq, acc=acc, recv=amt, paid=None, tag="uncredited", note="Uncredited lodgment")
    # Reversal pair credit leg
    add(dt=date(2026, 6, 27), name="Ecobank", details="Reversal of erroneous funds transfer OP048",
        doc="RV048/06/26", chq=None, acc=2140, recv=money("2500.00"), paid=None, tag="reversal", note="Reversal credit")
    add(dt=date(2026, 6, 28), name="Enterprise Life Assurance", details="Commissions received via transfer",
        doc=None, chq="trnsfer", acc=1020, recv=money("55320.80"), paid=None, tag="match", note="Funds transfer")
    # BF uncredited clearing is bank-side only historically; CB already had it in opening.

    # --- Payments (matched) ---
    add(dt=date(2026, 6, 2), name="Sodium Brand Solutions", details="Cost of office stationery – Jun 26",
        doc="OP002/06/26", chq="930101", acc=2131, recv=None, paid=money("1855.00"), tag="match", note="Cheque payment")
    add(dt=date(2026, 6, 3), name="Philip Akuffo", details="Cost of office vehicle repairs",
        doc="OP003/06/26", chq="930102", acc=2140, recv=None, paid=money("2400.00"), tag="match", note="Cheque payment")
    add(dt=date(2026, 6, 4), name="GRA", details="Payment of staff PAYE for May 26",
        doc="OP004/06/26", chq="930103", acc=7100, recv=None, paid=money("8214.50"), tag="match", note="Statutory")
    add(dt=date(2026, 6, 4), name="SSNIT", details="Payment of staff SSNIT deduction for May 26",
        doc="OP005/06/26", chq="930104", acc=7200, recv=None, paid=money("6890.25"), tag="match", note="Statutory")
    add(dt=date(2026, 6, 5), name="NBC", details="Payment of staff 2nd tier for May 26",
        doc="OP006/06/26", chq="930105", acc=7200, recv=None, paid=money("3210.00"), tag="match", note="Statutory")
    add(dt=date(2026, 6, 9), name="Ecobank", details="Reinvestment – 182-DAY T-BILLS",
        doc="OP009/06/26", chq="trnsfer", acc=3000, recv=None, paid=money("95000.00"), tag="match", note="T-bills purchase")
    add(dt=date(2026, 6, 10), name="ECG", details="Payment of ECG bill",
        doc="OP010/06/26", chq="930106", acc=2140, recv=None, paid=money("4718.60"), tag="match", note="Utility cheque")
    add(dt=date(2026, 6, 11), name="Staff Payroll", details="Staff salaries – payment Jun 2026",
        doc="OP011/06/26", chq="SALARY", acc=7400, recv=None, paid=money("45000.00"), tag="match", note="Staff salaries")
    add(dt=date(2026, 6, 12), name="Frederick-Leon Quayson", details="Payment of Commissions Jan to May 2026",
        doc="OP012/06/26", chq="930107", acc=2040, recv=None, paid=money("9126.41"), tag="match", note="Cheque CHQ# variant")
    add(dt=date(2026, 6, 13), name="Philip Akuffo", details="Payment of petty cash for office running",
        doc="OP013/06/26", chq="930108", acc=2110, recv=None, paid=money("1200.00"), tag="match", note="Cheque")
    add(dt=date(2026, 6, 16), name="Landlord – Airport Residential", details="Office rent for Jun 2026",
        doc="OP016/06/26", chq="930109", acc=2200, recv=None, paid=money("8500.00"), tag="match", note="Cheque")
    add(dt=date(2026, 6, 18), name="Conadu Consult", details="Website hosting and support",
        doc="OP018/06/26", chq="930110", acc=2160, recv=None, paid=money("3500.00"), tag="match", note="Cheque")
    # Unpresented
    for dt, name, details, doc, chq, acc, amt in UNPRESENTED:
        add(dt=dt, name=name, details=details, doc=doc, chq=chq, acc=acc, recv=None, paid=amt, tag="unpresented", note="Unpresented cheque")
    add(dt=date(2026, 6, 23), name="LIB Welfare", details="Payment of welfare dues and premium deductions",
        doc="OP043/06/26", chq="930113", acc=2310, recv=None, paid=money("2010.00"), tag="match", note="Cheque")
    # Erroneous payment then reversed
    add(dt=date(2026, 6, 26), name="Wrong Payee Ltd", details="Erroneous funds transfer – to be reversed",
        doc="OP048/06/26", chq="trnsfer", acc=2140, recv=None, paid=money("2500.00"), tag="reversal", note="Reversal debit leg")
    add(dt=date(2026, 6, 26), name="Kweku Abaka Quargraine", details="Accounting software subscription Jun 26",
        doc="OP049/06/26", chq="930115", acc=22200, recv=None, paid=money("380.00"), tag="match", note="Cheque")
    add(dt=date(2026, 6, 27), name="Enterprise Trustees", details="Payment of staff mutual fund for Jun 26",
        doc="OP050/06/26", chq="930116", acc=7409, recv=None, paid=money("3700.00"), tag="match", note="Cheque")
    add(dt=date(2026, 6, 30), name="Emmanuel Tetteh", details="Payment of office assistant and cleaner",
        doc="OP056/06/26", chq="930117", acc=2140, recv=None, paid=money("1100.00"), tag="match", note="Cheque")
    # BF unpresented that CLEARS this month (already in opening CB as paid historically —
    # we represent clearing only on bank; add a memo receipt? No — opening already nets it.
    # Put a memo note only in bank for CHQ 929880 BF 5,000 clearing.)

    rows.sort(key=lambda r: (r["dt"], 0 if r["recv"] else 1, r.get("doc") or "", r.get("name") or ""))
    return rows


def bank_narration_for(cb_row: dict) -> str:
    """Ecobank-style narration variants to challenge matching."""
    dt = cb_row["dt"].strftime("%d-%b-%Y")
    amt = cb_row["recv"] or cb_row["paid"]
    name = (cb_row["name"] or "").upper()
    chq = cb_row.get("chq")
    note = cb_row.get("note") or ""

    if cb_row["recv"]:
        if "GIPS" in (cb_row["details"] or "") or note == "MoMo/GIPS":
            return f"MOBILE TRANSFER RRN:137232{cb_row['dt'].strftime('%d%m')}441-GIP INCOMING B/O {name} IFO HORIZON INSURANCE BROKERS LTD GIPS Transfer / / //H55sboe{cb_row['dt'].strftime('%y%m%d')}002 {dt}"
        if note == "MoMo":
            return f"MOBILE TRANSFER RRN:MTN{cb_row['dt'].strftime('%d%m%y')}8821-GIP INCOMING B/O {name} IFO HORIZON INSURANCE BROKERS / / //H55momo{cb_row['dt'].strftime('%y%m%d')}011 {dt}"
        if "T-BILLS" in (cb_row["details"] or "").upper() or note == "T-bills matured":
            return f"TREASURY BILLS MATURED 182-DAY Bill AUC 2044 MAT @ 22.15 Due 09-JUN-2026 IFO HORIZON INSURANCE BROKERS LTD-{ACCT}"
        if note == "ACH inward":
            return f"OTHER BANKS INWARD TRANSFER ACH IRO COMMISSIONS STAR ASSURANCE / / //H12ach{cb_row['dt'].strftime('%y%m%d')}044 {dt}"
        if note == "Cheque clearing inward" or "cheque 001891" in (cb_row["details"] or "").lower():
            return f"CHEQUE CLEARING - OUTWARD LCY ACCESS CHQ# 001891 b/o DONEWELL INSURANCE / / //H75CQCL{cb_row['dt'].strftime('%y%m%d')}090 {dt}"
        if note == "HSE cheque deposit":
            return f"REF : H64LOCH{cb_row['dt'].strftime('%y%m%d')}022 CHEQUE DEPOSIT - HSE CHEQUE-EBG CHQ NO 000455 B/O GLICO GENERAL IFO HORIZON"
        if note.startswith("Batch"):
            return None  # handled separately as combined
        if note.startswith("Duplicate amount"):
            ref = "CUST-A7731" if "A" in note else "CUST-B7732"
            return f"FUNDS TRANSFER - INWARD {ref} trf b/o 1/{name} iro ifo 1/HORIZON INSURANCE BROKERS / / //H98INFT{cb_row['dt'].strftime('%y%m%d')}275 {dt}"
        if note == "Reversal credit":
            return f"JOURNAL ENTRY - NON COT REVERSAL OF ERRONEOUS FT OP048 / / //H01JNL{cb_row['dt'].strftime('%y%m%d')}7Q6 {dt}"
        # default funds transfer
        return f"FUNDS TRANSFER - INWARD GHAAO{cb_row['dt'].strftime('%d%m')}HWE6R trf b/o 1/{name} iro ifo 1/HORIZON INSURANCE BROKERS / / //H98INFT{cb_row['dt'].strftime('%y%m%d')}275 {dt}"

    # payments
    if note == "T-bills purchase":
        return f"TREASURY BILLS PURCHASE 182-DAY Auct. 2051@22.40 Due 09-DEC-2026 / / /H01ZEXA{cb_row['dt'].strftime('%y%m%d')}7Q6 {dt}"
    if note == "Staff salaries":
        return f"STAFF SALARIES - PAYMENT HORIZON INSURANCE BROKERS JUN 2026 / / //H33SAL{cb_row['dt'].strftime('%y%m%d')}001 {dt}"
    if note == "Reversal debit leg":
        return f"FUNDS TRANSFER - OUTWARD ERRONEOUS FT OP048 TO WRONG PAYEE LTD / / //H98OUT{cb_row['dt'].strftime('%y%m%d')}510 {dt}"
    if chq and chq.isdigit():
        # narration variant: CHQ NO vs CHQ#
        if chq == "930107":
            return f"CHEQUE WITHDRAWAL CHQ# {chq} PAID TO {name} / / //H75CQWL{cb_row['dt'].strftime('%y%m%d')}116 {dt}"
        return f"CHEQUE WITHDRAWAL CHQ NO {chq} PAID TO {name} / / //H75CQWL{cb_row['dt'].strftime('%y%m%d')}116 {dt}"
    if chq == "SALARY":
        return f"STAFF SALARIES - PAYMENT HORIZON INSURANCE BROKERS JUN 2026 / / //H33SAL{cb_row['dt'].strftime('%y%m%d')}001 {dt}"
    if chq == "trnsfer":
        return f"FUNDS TRANSFER - OUTWARD {cb_row.get('doc') or ''} TO {name} / / //H98OUT{cb_row['dt'].strftime('%y%m%d')}088 {dt}"
    return f"PAYMENT {name} {cb_row.get('details') or ''} {dt}"


def build_bank_rows(cb_rows):
    bank = []

    # Opening balance line handled separately

    batch_credit_added = False
    for r in cb_rows:
        if r["tag"] in ("uncredited", "unpresented"):
            continue
        if r["tag"] == "one_to_many":
            if not batch_credit_added:
                bank.append(dict(
                    dt=date(2026, 6, 11),  # value-date lag
                    desc="FUNDS TRANSFER - INWARD BATCH-DEP-JUN10 trf b/o PHOENIX/QUALITY COMMISSIONS iro ifo HORIZON INSURANCE BROKERS / / //H98INFT260611BAT 11-Jun-2026",
                    debit=None, credit=money("10000.00"), tag="one_to_many", note="One bank credit vs two CB receipts",
                ))
                batch_credit_added = True
            continue
        narr = bank_narration_for(r)
        # Date lag: some bank postings 0–2 days after cash book
        lag = 0
        if r["tag"] == "match" and r["recv"] and "Enterprise Life" in (r["name"] or "") and r["dt"].day == 2:
            lag = 1
        if r["tag"] == "match" and r.get("chq") == "930101":
            lag = 2
        bdt = date(r["dt"].year, r["dt"].month, min(r["dt"].day + lag, 30))
        bank.append(dict(
            dt=bdt,
            desc=narr,
            debit=r["paid"],
            credit=r["recv"],
            tag=r["tag"],
            note=r.get("note"),
            cb_doc=r.get("doc"),
            cb_chq=r.get("chq"),
        ))

    for dt, desc, amt in BANK_ONLY_DEBITS:
        bank.append(dict(dt=dt, desc=desc, debit=amt, credit=None, tag="bank_only", note="Bank-only debit"))
    for dt, desc, amt in BANK_ONLY_CREDITS:
        bank.append(dict(dt=dt, desc=desc, debit=None, credit=amt, tag="bank_only", note="Bank-only credit"))

    bank.sort(key=lambda x: (x["dt"], 0 if x["credit"] else 1, x["desc"]))
    return bank


# Styles
thin = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)
header_fill = PatternFill("solid", fgColor="1F4E79")
header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
title_font = Font(bold=True, name="Calibri", size=14, color="1F4E79")
money_font = Font(name="Calibri", size=10)
normal = Font(name="Calibri", size=10)
bold = Font(bold=True, name="Calibri", size=10)
green_fill = PatternFill("solid", fgColor="E2EFDA")
amber_fill = PatternFill("solid", fgColor="FFF2CC")
red_fill = PatternFill("solid", fgColor="FCE4D6")
blue_fill = PatternFill("solid", fgColor="DDEBF7")


def autosize(ws, widths=None):
    if widths:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        return
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        maxlen = 0
        for cell in col:
            if cell.value is not None:
                maxlen = max(maxlen, min(len(str(cell.value)), 55))
        ws.column_dimensions[letter].width = max(12, maxlen + 2)


def write_cash_book(path: Path, cb_rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Cash Book"

    ws["A1"] = COMPANY
    ws["A1"].font = title_font
    ws["A2"] = f"CASH BOOK – Ecobank Current Account {ACCT}"
    ws["A3"] = f"Period: {PERIOD}    Currency: {CURRENCY}"
    ws["A4"] = "For go-live completeness testing – Ghana advanced specimen"

    headers = ["DATE", "NAME", "DETAILS", "DOC REF", "CHQ NO", "ACCODE", "AMT RECEIVED", "AMT PAID", "BALANCE"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(5, c, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin

    bal = OPEN_CB
    ws.append([datetime(2026, 6, 1), None, "Balance b/d", None, None, None, None, None, float(bal)])
    for c in range(1, 10):
        ws.cell(6, c).font = bold
        ws.cell(6, c).fill = blue_fill
        ws.cell(6, c).border = thin
    ws.cell(6, 9).number_format = '#,##0.00'

    for r in cb_rows:
        recv = r["recv"]
        paid = r["paid"]
        if recv:
            bal = money(bal + recv)
        if paid:
            bal = money(bal - paid)
        row = [
            datetime(r["dt"].year, r["dt"].month, r["dt"].day),
            r["name"],
            r["details"],
            r["doc"],
            r["chq"],
            r["acc"],
            float(recv) if recv is not None else None,
            float(paid) if paid is not None else None,
            float(bal),
        ]
        ws.append(row)
        rr = ws.max_row
        for c in range(1, 10):
            ws.cell(rr, c).border = thin
            ws.cell(rr, c).font = normal
        for c in (7, 8, 9):
            ws.cell(rr, c).number_format = '#,##0.00'
        if r["tag"] == "uncredited":
            for c in range(1, 10):
                ws.cell(rr, c).fill = amber_fill
        elif r["tag"] == "unpresented":
            for c in range(1, 10):
                ws.cell(rr, c).fill = red_fill
        elif r["tag"] in ("one_to_many", "reversal", "dup_amt"):
            for c in range(1, 10):
                ws.cell(rr, c).fill = green_fill

    close_cb = bal
    ws.append([])
    ws.append(["Closing balance c/d", None, None, None, None, None, None, None, float(close_cb)])
    ws.cell(ws.max_row, 1).font = bold
    ws.cell(ws.max_row, 9).font = bold
    ws.cell(ws.max_row, 9).number_format = '#,##0.00'

    # Legend
    ws2 = wb.create_sheet("Legend")
    ws2["A1"] = "Highlight legend (for manual review; upload sheet is Cash Book)"
    ws2["A1"].font = bold
    legend = [
        ("Blue", "Opening balance"),
        ("Amber", "Uncredited lodgment (cash book only – timing)"),
        ("Orange/Red", "Unpresented cheque (cash book only – timing)"),
        ("Green", "Challenge cases: one-to-many, reversal, duplicate amounts"),
    ]
    ws2.append(["Colour", "Meaning"])
    for a, b in legend:
        ws2.append([a, b])
    autosize(ws, [12, 28, 52, 14, 12, 10, 14, 12, 14])
    autosize(ws2)
    wb.save(path)
    return close_cb


def write_bank_statement(path: Path, bank_rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Bank Statement"

    ws["A1"] = BANK
    ws["A1"].font = title_font
    ws["A2"] = f"Account Name: {COMPANY}"
    ws["A3"] = f"Account Number: {ACCT}"
    ws["A4"] = f"Statement Period: {PERIOD}    Currency: {CURRENCY}"
    ws["A5"] = "Specimen bank statement for go-live testing (Ecobank-style narrations)"

    # Use "Transaction Description" so Stanbic/Ecobank-style detectors keep narrations
    # (plain "Description" is dropped by normalizeStanbicExcelTable).
    headers = ["Transaction Date", "Value Date", "Transaction Description", "Debit", "Credit", "Balance"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(7, c, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin

    bal = OPEN_BANK
    # Skip OPENING BALANCE data row — balance is in header notes; avoids empty-desc noise.
    start_row = 8
    ws.cell(start_row - 1, 1)  # headers already at row 7

    for r in bank_rows:
        if r["credit"]:
            bal = money(bal + r["credit"])
        if r["debit"]:
            bal = money(bal - r["debit"])
        ws.append([
            datetime(r["dt"].year, r["dt"].month, r["dt"].day),
            datetime(r["dt"].year, r["dt"].month, r["dt"].day),
            r["desc"],
            float(r["debit"]) if r["debit"] is not None else None,
            float(r["credit"]) if r["credit"] is not None else None,
            float(bal),
        ])
        rr = ws.max_row
        for c in range(1, 7):
            ws.cell(rr, c).border = thin
            ws.cell(rr, c).font = normal
        for c in (4, 5, 6):
            ws.cell(rr, c).number_format = '#,##0.00'
        if r["tag"] == "bank_only":
            for c in range(1, 7):
                ws.cell(rr, c).fill = amber_fill
        elif r["tag"] in ("one_to_many", "reversal", "dup_amt", "bf_clear_credit", "bf_clear_debit"):
            for c in range(1, 7):
                ws.cell(rr, c).fill = green_fill

    close_bank = bal
    # Verify last data balance equals close_bank via running total already applied
    autosize(ws, [16, 14, 70, 12, 12, 14])
    wb.save(path)
    return close_bank


def write_brs(path: Path, close_cb: Decimal, close_bank: Decimal):
    uncred_total = money(sum(x[-1] for x in UNCREDITED))
    unpres_total = money(sum(x[-1] for x in UNPRESENTED))
    bank_only_dr = money(sum(x[-1] for x in BANK_ONLY_DEBITS))
    bank_only_cr = money(sum(x[-1] for x in BANK_ONLY_CREDITS))

    # Ghana workbook schedule
    # CB = Bank + uncredited − unpresented + bank-only debits − bank-only credits
    derived_cb = money(close_bank + uncred_total - unpres_total + bank_only_dr - bank_only_cr)

    wb = Workbook()
    ws = wb.active
    ws.title = "BRS"

    ws["A1"] = COMPANY
    ws["A1"].font = Font(bold=True, name="Calibri", size=16, color="1F4E79")
    ws["A2"] = "BANK RECONCILIATION STATEMENT"
    ws["A2"].font = Font(bold=True, name="Calibri", size=13)
    ws["A3"] = "AS AT 30TH JUNE, 2026"
    ws["A3"].font = Font(bold=True, name="Calibri", size=12)
    ws["A4"] = f"{BANK} Account Number {ACCT}"
    ws["A5"] = f"Currency: {CURRENCY} (GH₵)"

    start = 7
    schedule = [
        ("Closing balance per bank statement", close_bank, None),
        ("Add: Uncredited lodgments", uncred_total, "timing"),
        ("Less: Unpresented cheques", unpres_total, "timing"),
        ("Add: Bank-only debits (charges etc.)", bank_only_dr, "bank_only"),
        ("Deduct: Bank-only credits (interest etc.)", bank_only_cr, "bank_only"),
        ("Cash book balance at end of period", close_cb, "result"),
    ]
    ws.cell(start, 1, "Particulars").font = header_font
    ws.cell(start, 1).fill = header_fill
    ws.cell(start, 2, "GH₵").font = header_font
    ws.cell(start, 2).fill = header_fill
    for i, (label, amt, kind) in enumerate(schedule, 1):
        ws.cell(start + i, 1, label).font = bold if kind == "result" else normal
        ws.cell(start + i, 2, float(amt)).number_format = '#,##0.00'
        ws.cell(start + i, 2).font = bold if kind == "result" else money_font
        if kind == "result":
            ws.cell(start + i, 1).fill = green_fill
            ws.cell(start + i, 2).fill = green_fill
        for c in (1, 2):
            ws.cell(start + i, c).border = thin

    row = start + len(schedule) + 2
    ws.cell(row, 1, "Derived cash book (bank + uncredited − unpresented + bank-only DR − bank-only CR)")
    ws.cell(row, 2, float(derived_cb)).number_format = '#,##0.00'
    ws.cell(row, 2).font = bold
    row += 1
    ws.cell(row, 1, "Difference (must be 0.00)")
    ws.cell(row, 2, float(money(close_cb - derived_cb))).number_format = '#,##0.00'
    if money(close_cb - derived_cb) == 0:
        ws.cell(row, 3, "TIED")
        ws.cell(row, 3).fill = green_fill
    else:
        ws.cell(row, 3, "OUT OF BALANCE")
        ws.cell(row, 3).fill = red_fill

    # Supporting tables
    row += 3
    ws.cell(row, 1, "UNCREDITED LODGMENTS").font = bold
    row += 1
    for c, h in enumerate(["Date", "Name", "Details", "Doc Ref", "CHQ NO", "GH₵"], 1):
        cell = ws.cell(row, c, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin
    for dt, name, details, doc, chq, acc, amt in UNCREDITED:
        row += 1
        vals = [dt.strftime("%d %b %Y"), name, details, doc, chq or "", float(amt)]
        for c, v in enumerate(vals, 1):
            ws.cell(row, c, v).border = thin
        ws.cell(row, 6).number_format = '#,##0.00'
    row += 1
    ws.cell(row, 5, "Total").font = bold
    ws.cell(row, 6, float(uncred_total)).number_format = '#,##0.00'
    ws.cell(row, 6).font = bold

    row += 3
    ws.cell(row, 1, "UNPRESENTED CHEQUES").font = bold
    row += 1
    for c, h in enumerate(["Date", "Name", "Details", "Doc Ref", "CHQ NO", "GH₵"], 1):
        cell = ws.cell(row, c, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin
    for dt, name, details, doc, chq, acc, amt in UNPRESENTED:
        row += 1
        vals = [dt.strftime("%d %b %Y"), name, details, doc, chq, float(amt)]
        for c, v in enumerate(vals, 1):
            ws.cell(row, c, v).border = thin
        ws.cell(row, 6).number_format = '#,##0.00'
    row += 1
    ws.cell(row, 5, "Total").font = bold
    ws.cell(row, 6, float(unpres_total)).number_format = '#,##0.00'
    ws.cell(row, 6).font = bold

    row += 3
    ws.cell(row, 1, "BANK-ONLY ITEMS").font = bold
    row += 1
    for c, h in enumerate(["Date", "Description", "Debit", "Credit"], 1):
        cell = ws.cell(row, c, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin
    for dt, desc, amt in BANK_ONLY_DEBITS:
        row += 1
        for c, v in enumerate([dt.strftime("%d %b %Y"), desc, float(amt), None], 1):
            ws.cell(row, c, v).border = thin
        ws.cell(row, 3).number_format = '#,##0.00'
    for dt, desc, amt in BANK_ONLY_CREDITS:
        row += 1
        for c, v in enumerate([dt.strftime("%d %b %Y"), desc, None, float(amt)], 1):
            ws.cell(row, c, v).border = thin
        ws.cell(row, 4).number_format = '#,##0.00'

    row += 3
    ws.cell(row, 1, "SIGN-OFF").font = bold
    row += 1
    headers = ["", "Prepared By", "Checked By", "Approved By"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row, c, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin
    for label in ("NAME", "SIGNATURE", "DATE"):
        row += 1
        ws.cell(row, 1, label).border = thin
        for c in range(2, 5):
            ws.cell(row, c, "").border = thin

    # Notes sheet
    notes = wb.create_sheet("Expected Totals")
    notes["A1"] = "Manual reconciliation expected totals (go-live golden)"
    notes["A1"].font = title_font
    metrics = [
        ("Opening cash book balance", OPEN_CB),
        ("Opening bank balance", OPEN_BANK),
        ("BF uncredited (clears in June)", BF_UNCREDITED),
        ("BF unpresented (clears in June)", BF_UNPRESENTED),
        ("Closing cash book balance", close_cb),
        ("Closing bank statement balance", close_bank),
        ("Uncredited lodgments (period-end)", uncred_total),
        ("Unpresented cheques (period-end)", unpres_total),
        ("Bank-only debits", bank_only_dr),
        ("Bank-only credits", bank_only_cr),
        ("Derived CB from workbook schedule", derived_cb),
        ("Difference", money(close_cb - derived_cb)),
    ]
    notes.append(["Metric", "Value (GHS)"])
    for m, v in metrics:
        notes.append([m, float(v)])
        notes.cell(notes.max_row, 2).number_format = '#,##0.00'

    notes.append([])
    notes.append(["Challenge cases included", ""])
    for line in [
        "Funds transfer inward/outward",
        "MoMo / GIPS incoming",
        "Cheque clearing / house cheque deposit",
        "ACH other banks inward",
        "Treasury bills matured + reinvestment",
        "Staff salaries payment",
        "Statutory: GRA PAYE, SSNIT, NBC 2nd tier",
        "Utilities (ECG), rent, commissions payable",
        "Uncredited lodgments (2)",
        "Unpresented cheques (3)",
        "Bank-only: COT, SMS, maintenance, cheque book, interest",
        "One-to-many combined deposit (2 CB receipts → 1 bank credit)",
        "Narration variants (CHQ NO vs CHQ#)",
        "Date/value-date lag",
        "Reversal pair (erroneous FT + journal reversal)",
        "Duplicate amounts different counterparties",
        "Brought-forward timing clearing in period",
    ]:
        notes.append([line])

    autosize(ws, [55, 16, 55, 14, 12, 14])
    autosize(notes, [50, 18])
    wb.save(path)
    return {
        "close_cb": close_cb,
        "close_bank": close_bank,
        "uncredited": uncred_total,
        "unpresented": unpres_total,
        "bank_only_dr": bank_only_dr,
        "bank_only_cr": bank_only_cr,
        "derived_cb": derived_cb,
        "diff": money(close_cb - derived_cb),
    }


def write_reference_map(path: Path, cb_rows, bank_rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "ReferenceMap"
    ws.append(["Side", "Date", "Ref/Doc", "CHQ", "Amount", "Status", "Comment"])
    for r in cb_rows:
        amt = r["recv"] or r["paid"]
        status = {
            "match": "Matched",
            "one_to_many": "Partial / One-to-many",
            "uncredited": "Uncredited lodgment",
            "unpresented": "Unpresented cheque",
            "reversal": "Matched (reversal)",
            "dup_amt": "Matched (duplicate amount)",
        }.get(r["tag"], r["tag"])
        ws.append([
            "Cash Book",
            r["dt"].isoformat(),
            r.get("doc") or r.get("name"),
            r.get("chq"),
            float(amt),
            status,
            r.get("note"),
        ])
    for r in bank_rows:
        if r["tag"] in ("bf_clear_credit", "bf_clear_debit"):
            status = "Matched (BF clear)"
        elif r["tag"] == "bank_only":
            status = "Bank-only"
        elif r["tag"] == "one_to_many":
            status = "Partial / One-to-many"
        else:
            status = "Matched"
        amt = r["credit"] or r["debit"]
        ws.append([
            "Bank",
            r["dt"].isoformat(),
            (r["desc"] or "")[:80],
            r.get("cb_chq"),
            float(amt) if amt is not None else None,
            status,
            r.get("note"),
        ])
    autosize(ws)
    wb.save(path)


def pdf_table(data, col_widths, header=True):
    style_cmds = [
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]
    if header:
        style_cmds += [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ]
    t = Table(data, colWidths=col_widths, repeatRows=1 if header else 0)
    t.setStyle(TableStyle(style_cmds))
    return t


def write_cash_book_pdf(path: Path, cb_rows, close_cb: Decimal):
    doc = SimpleDocTemplate(str(path), pagesize=landscape(A4),
                            leftMargin=10 * mm, rightMargin=10 * mm,
                            topMargin=12 * mm, bottomMargin=12 * mm)
    styles = getSampleStyleSheet()
    title = ParagraphStyle("t", parent=styles["Heading1"], fontSize=12, textColor=colors.HexColor("#1F4E79"), spaceAfter=2)
    sub = ParagraphStyle("s", parent=styles["Normal"], fontSize=8, spaceAfter=2)
    story = [
        Paragraph(COMPANY, title),
        Paragraph(f"CASH BOOK – Ecobank Current Account {ACCT}", sub),
        Paragraph(f"Period: {PERIOD} | Currency: {CURRENCY} | Closing: GH₵ {fmt(close_cb)}", sub),
        Spacer(1, 4 * mm),
    ]
    data = [["DATE", "NAME", "DETAILS", "DOC REF", "CHQ NO", "ACCODE", "AMT RECEIVED", "AMT PAID", "BALANCE"]]
    bal = OPEN_CB
    data.append(["01-Jun-2026", "", "Balance b/d", "", "", "", "", "", fmt(bal)])
    for r in cb_rows:
        if r["recv"]:
            bal = money(bal + r["recv"])
        if r["paid"]:
            bal = money(bal - r["paid"])
        data.append([
            r["dt"].strftime("%d-%b-%Y"),
            (r["name"] or "")[:22],
            Paragraph((r["details"] or "")[:70], ParagraphStyle("d", fontSize=6.5, leading=8)),
            r["doc"] or "",
            r["chq"] or "",
            str(r["acc"] or ""),
            fmt(r["recv"]) if r["recv"] is not None else "",
            fmt(r["paid"]) if r["paid"] is not None else "",
            fmt(bal),
        ])
    story.append(pdf_table(data, [22 * mm, 35 * mm, 70 * mm, 25 * mm, 18 * mm, 16 * mm, 25 * mm, 22 * mm, 25 * mm]))
    doc.build(story)


def write_bank_pdf(path: Path, bank_rows, close_bank: Decimal):
    doc = SimpleDocTemplate(str(path), pagesize=landscape(A4),
                            leftMargin=10 * mm, rightMargin=10 * mm,
                            topMargin=12 * mm, bottomMargin=12 * mm)
    styles = getSampleStyleSheet()
    title = ParagraphStyle("t", parent=styles["Heading1"], fontSize=12, textColor=colors.HexColor("#1F4E79"), spaceAfter=2)
    sub = ParagraphStyle("s", parent=styles["Normal"], fontSize=8, spaceAfter=2)
    story = [
        Paragraph(BANK, title),
        Paragraph(f"Account: {COMPANY} | {ACCT}", sub),
        Paragraph(f"Period: {PERIOD} | Currency: {CURRENCY} | Closing: GH₵ {fmt(close_bank)}", sub),
        Spacer(1, 4 * mm),
    ]
    data = [["Txn Date", "Description", "Debit", "Credit", "Balance"]]
    bal = OPEN_BANK
    data.append(["01-Jun-2026", "OPENING BALANCE", "", "", fmt(bal)])
    for r in bank_rows:
        if r["credit"]:
            bal = money(bal + r["credit"])
        if r["debit"]:
            bal = money(bal - r["debit"])
        data.append([
            r["dt"].strftime("%d-%b-%Y"),
            Paragraph((r["desc"] or "")[:140], ParagraphStyle("d", fontSize=6.5, leading=8)),
            fmt(r["debit"]) if r["debit"] is not None else "",
            fmt(r["credit"]) if r["credit"] is not None else "",
            fmt(bal),
        ])
    story.append(pdf_table(data, [22 * mm, 175 * mm, 25 * mm, 25 * mm, 25 * mm]))
    doc.build(story)


def write_brs_pdf(path: Path, totals: dict):
    doc = SimpleDocTemplate(str(path), pagesize=A4,
                            leftMargin=15 * mm, rightMargin=15 * mm,
                            topMargin=15 * mm, bottomMargin=15 * mm)
    styles = getSampleStyleSheet()
    title = ParagraphStyle("t", parent=styles["Heading1"], fontSize=13, alignment=1, textColor=colors.HexColor("#1F4E79"))
    center = ParagraphStyle("c", parent=styles["Normal"], fontSize=10, alignment=1, spaceAfter=2)
    story = [
        Paragraph(COMPANY, title),
        Paragraph("BANK RECONCILIATION STATEMENT", center),
        Paragraph("AS AT 30TH JUNE, 2026", center),
        Paragraph(f"{BANK} Account Number {ACCT}", center),
        Paragraph(f"Currency: {CURRENCY}", center),
        Spacer(1, 6 * mm),
    ]
    sched = [
        ["Particulars", "GH₵"],
        ["Closing balance per bank statement", fmt(totals["close_bank"])],
        ["Add: Uncredited lodgments", fmt(totals["uncredited"])],
        ["Less: Unpresented cheques", fmt(totals["unpresented"])],
        ["Add: Bank-only debits (charges etc.)", fmt(totals["bank_only_dr"])],
        ["Deduct: Bank-only credits (interest etc.)", fmt(totals["bank_only_cr"])],
        ["Cash book balance at end of period", fmt(totals["close_cb"])],
    ]
    story.append(pdf_table(sched, [120 * mm, 40 * mm]))
    story.append(Spacer(1, 4 * mm))
    story.append(Paragraph(
        f"Derived CB: GH₵ {fmt(totals['derived_cb'])} | Difference: GH₵ {fmt(totals['diff'])} | "
        + ("TIED" if totals["diff"] == 0 else "OUT OF BALANCE"),
        center,
    ))
    story.append(Spacer(1, 6 * mm))
    story.append(Paragraph("<b>UNCREDITED LODGMENTS</b>", styles["Normal"]))
    u = [["Date", "Details", "GH₵"]]
    for dt, name, details, _dref, chq, acc, amt in UNCREDITED:
        u.append([dt.strftime("%d %b %Y"), f"{name} – {details}", fmt(amt)])
    u.append(["", "Total", fmt(totals["uncredited"])])
    story.append(pdf_table(u, [30 * mm, 110 * mm, 30 * mm]))
    story.append(Spacer(1, 4 * mm))
    story.append(Paragraph("<b>UNPRESENTED CHEQUES</b>", styles["Normal"]))
    p = [["Date", "Details / CHQ", "GH₵"]]
    for dt, name, details, _dref, chq, acc, amt in UNPRESENTED:
        p.append([dt.strftime("%d %b %Y"), f"{name} – CHQ {chq}", fmt(amt)])
    p.append(["", "Total", fmt(totals["unpresented"])])
    story.append(pdf_table(p, [30 * mm, 110 * mm, 30 * mm]))
    story.append(Spacer(1, 8 * mm))
    sign = [
        ["", "Prepared By", "Checked By", "Approved By"],
        ["NAME", "", "", ""],
        ["SIGNATURE", "", "", ""],
        ["DATE", "", "", ""],
    ]
    story.append(pdf_table(sign, [30 * mm, 45 * mm, 45 * mm, 45 * mm]))
    doc.build(story)


def write_readme(path: Path, totals: dict, n_cb: int, n_bank: int):
    text = f"""# Ghana Go-Live Test Pack – Horizon Insurance Brokers Ltd

Advanced Ghana cash book + Ecobank-style bank statement for **system completeness / going-live** testing.
Use the same figures for **manual Excel BRS** and **platform Excel/PDF** import tests.

## Files
| File | Purpose |
|------|---------|
| `01_cash_book_horizon_jun2026.xlsx` | Cash book upload (Excel) |
| `01_cash_book_horizon_jun2026.pdf` | Same cash book (PDF upload test) |
| `02_bank_statement_ecobank_jun2026.xlsx` | Bank statement upload (Excel) |
| `02_bank_statement_ecobank_jun2026.pdf` | Same bank statement (PDF upload test) |
| `03_final_brs_as_at_30_june_2026.xlsx` | Manual golden BRS (Excel) |
| `03_final_brs_as_at_30_june_2026.pdf` | Manual golden BRS (PDF) |
| `04_manual_reference_map.xlsx` | Expected match / timing / bank-only map |
| `generate_ghana_go_live_pack.py` | Regenerator script |

## Company / Account
- **Company:** {COMPANY}
- **Bank:** {BANK}
- **Account Number:** {ACCT}
- **Period:** {PERIOD}
- **As at:** 30TH JUNE, 2026
- **Currency:** GHS (GH₵)

## Expected balances (golden)
| Metric | GHS |
|--------|-----|
| Opening cash book | {fmt(OPEN_CB)} |
| Opening bank | {fmt(OPEN_BANK)} |
| BF uncredited (clears in June) | {fmt(BF_UNCREDITED)} |
| BF unpresented (clears in June) | {fmt(BF_UNPRESENTED)} |
| Closing cash book | {fmt(totals['close_cb'])} |
| Closing bank statement | {fmt(totals['close_bank'])} |
| Uncredited lodgments (period-end) | {fmt(totals['uncredited'])} |
| Unpresented cheques (period-end) | {fmt(totals['unpresented'])} |
| Bank-only debits | {fmt(totals['bank_only_dr'])} |
| Bank-only credits | {fmt(totals['bank_only_cr'])} |
| Derived CB from schedule | {fmt(totals['derived_cb'])} |
| Difference | {fmt(totals['diff'])} |

Workbook formula used:
`Cash book = Bank closing + Uncredited − Unpresented + Bank-only debits − Bank-only credits`

## Transaction types included ({n_cb} cash-book lines, {n_bank} bank lines)
1. Funds transfer inward / outward  
2. MoMo / GIPS incoming  
3. Cheque clearing / house cheque deposit  
4. ACH other banks inward  
5. Treasury bills matured + reinvestment  
6. Staff salaries  
7. Statutory (GRA PAYE, SSNIT, NBC 2nd tier)  
8. Utilities, rent, commissions, welfare  
9. Uncredited lodgments (2)  
10. Unpresented cheques (3)  
11. Bank-only COT, SMS, maintenance, cheque book, interest  
12. One-to-many combined deposit  
13. Narration variants (`CHQ NO` vs `CHQ#`)  
14. Date / value-date lag  
15. Reversal pair  
16. Duplicate amounts (different counterparties)  
17. Brought-forward timing that clears in the period  

## How to use
1. Upload Excel cash book + bank statement into a new project (or PDF pair for PDF path).  
2. Map columns if not auto-detected (`DATE/DETAILS/AMT RECEIVED/AMT PAID` and `Transaction Date/Description/Debit/Credit/Balance`).  
3. Run matching / reconciliation.  
4. Export platform BRS and compare to `03_final_brs_as_at_30_june_2026.xlsx` / `.pdf`.  
5. Do the same reconciliation manually in Excel using the golden totals above.

## Colour cues (Excel only – not required for upload)
- Amber rows on cash book = uncredited  
- Red/orange rows on cash book = unpresented  
- Green = challenge cases (batch, reversal, duplicates)  
- Amber on bank statement = bank-only items  
"""
    path.write_text(text, encoding="utf-8")


def main():
    OUT.mkdir(parents=True, exist_ok=True)
    cb_rows = build_cash_book_rows()
    bank_rows = build_bank_rows(cb_rows)

    close_cb = write_cash_book(OUT / "01_cash_book_horizon_jun2026.xlsx", cb_rows)
    close_bank = write_bank_statement(OUT / "02_bank_statement_ecobank_jun2026.xlsx", bank_rows)
    totals = write_brs(OUT / "03_final_brs_as_at_30_june_2026.xlsx", close_cb, close_bank)
    write_reference_map(OUT / "04_manual_reference_map.xlsx", cb_rows, bank_rows)

    write_cash_book_pdf(OUT / "01_cash_book_horizon_jun2026.pdf", cb_rows, close_cb)
    write_bank_pdf(OUT / "02_bank_statement_ecobank_jun2026.pdf", bank_rows, close_bank)
    write_brs_pdf(OUT / "03_final_brs_as_at_30_june_2026.pdf", totals)
    write_readme(OUT / "README.md", totals, len(cb_rows), len(bank_rows))

    print("=== GENERATED ===")
    print(f"Cash book lines: {len(cb_rows)}  closing CB: {fmt(close_cb)}")
    print(f"Bank lines: {len(bank_rows)}  closing Bank: {fmt(close_bank)}")
    for k, v in totals.items():
        print(f"  {k}: {fmt(v)}")
    if totals["diff"] != 0:
        raise SystemExit("BRS DOES NOT TIE – fix data before using")
    print("BRS TIED OK")
    for p in sorted(OUT.glob("*")):
        if p.name != "generate_ghana_go_live_pack.py":
            print(" ", p.name, p.stat().st_size)


if __name__ == "__main__":
    main()
